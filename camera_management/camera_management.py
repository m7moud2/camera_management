import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from datetime import datetime
import json
import os
from PIL import Image, ImageTk, ImageGrab
import io
import base64
import hashlib
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import threading
import time

# ==================== COLORS ====================
COLORS = {
    'bg_main': '#171717',           # gray_ultradark
    'bg_sidebar': '#282829',        # gray_dark
    'bg_card': '#282829',           # gray_dark
    'bg_input': '#171717',          # gray_ultradark
    'text_primary': '#ffffff',      # white
    'text_secondary': '#d5d6d6',    # gray_ultralight
    'text_dim': '#939598',          # gray_light
    'text_label': '#636466',        # gray
    'accent_blue': '#2e69ff',       # blue
    'accent_blue_dark': '#195faa',  # blue_dark
    'accent_green': '#2d9655',      # green
    'accent_red': '#b92b27',        # red
    'border': '#636466',            # gray
    'hover': '#3a3a3b'
}

# ==================== TRANSLATIONS ====================
TRANSLATIONS = {
    'ar': {
        'app_title': 'لوحة التحكم - إدارة الكاميرات',
        'login_title': 'تسجيل الدخول',
        'register_title': 'إنشاء حساب جديد',
        'username': 'اسم المستخدم',
        'password': 'كلمة المرور',
        'confirm_password': 'تأكيد كلمة المرور',
        'login': 'دخول',
        'register': 'تسجيل',
        'back_to_login': 'العودة لتسجيل الدخول',
        'logout': 'تسجيل خروج',
        'dashboard': 'لوحة التحكم',
        'cameras': 'الكاميرات',
        'add_camera': 'إضافة كاميرا',
        'statistics': 'الإحصائيات',
        'add': 'إضافة',
        'update': 'تحديث',
        'delete': 'حذف',
        'clear': 'مسح',
        'search': 'بحث',
        'export_excel': 'تصدير Excel',
        'import_excel': 'استيراد Excel',
        'take_screenshot': 'رفع صورة',
        'upload_image': 'اختر صورة من الجهاز',
        'camera_number': 'رقم الكاميرا',
        'branch_name': 'اسم الفرع',
        'branch_code': 'رمز الفرع',
        'device_ip': 'عنوان IP الجهاز',
        'dvr_name': 'اسم الـ DVR',
        'username_field': 'اسم المستخدم',
        'password_field': 'كلمة المرور',
        'location': 'الموقع التفصيلي',
        'location_type': 'نوع الموقع',
        'direction': 'الاتجاه',
        'working_hours': 'فترة العمل',
        'model': 'الموديل',
        'serial_number': 'الرقم التسلسلي',
        'resolution': 'الدقة',
        'features': 'المميزات',
        'select_features': 'اختر المميزات',
        'night_vision': 'رؤية ليلية',
        'motion_detection': 'كشف حركة',
        'audio': 'صوت',
        'ptz': 'PTZ (تحريك)',
        'waterproof': 'مقاوم للماء',
        'vandal_proof': 'مقاوم للتخريب',
        'face_recognition': 'تعرف على الوجوه',
        'license_plate': 'تعرف لوحات السيارات',
        'wide_angle': 'زاوية واسعة',
        'zoom': 'تقريب بصري',
        'two_way_audio': 'صوت ثنائي الاتجاه',
        'smart_tracking': 'تتبع ذكي',
        'lens_layout': 'نوع العدسة',
        'rtsp_url': 'رابط RTSP',
        'install_date': 'تاريخ التركيب',
        'warranty_expiry': 'تاريخ انتهاء الضمان',
        'recording_capacity': 'سعة التسجيل',
        'status': 'الحالة',
        'notes': 'ملاحظات',
        'screenshot': 'لقطة الشاشة',
        'ready': 'جاهز',
        'total_cameras': 'إجمالي الكاميرات',
        'active_cameras': 'كاميرات نشطة',
        'inactive_cameras': 'كاميرات غير نشطة',
        'maintenance_cameras': 'كاميرات صيانة',
        'recent_cameras': 'أحدث الكاميرات',
        'view_all': 'عرض الكل',
        'export_pdf': 'تصدير PDF',
        'print_report': 'طباعة تقرير',
        'filter': 'تصفية',
        'all_status': 'كل الحالات',
        'all_locations': 'كل المواقع',
        'apply_filter': 'تطبيق الفلتر',
        'reset_filter': 'إعادة تعيين',
        'camera_details': 'تفاصيل الكاميرا',
        'technical_info': 'معلومات تقنية',
        'system_info': 'معلومات النظام',
        'language': 'اللغة',
        'users_management': 'إدارة المستخدمين',
        'add_user': 'إضافة مستخدم',
        'edit_user': 'تعديل مستخدم',
        'delete_user': 'حذف مستخدم',
        'role': 'الصلاحية',
        'admin': 'مدير',
        'moderator': 'مشرف',
        'user': 'مستخدم',
        'created_by': 'أنشئ بواسطة',
        'created_at': 'تاريخ الإنشاء',
        'change_password': 'تغيير كلمة المرور',
        'new_password': 'كلمة المرور الجديدة',
        'user_status': 'حالة المستخدم',
        'active_user': 'نشط',
        'inactive_user': 'غير نشط',
        'vault': 'خزنة',
        'teller': 'تيللر',
        'customer_service': 'خدمة عملاء',
        'entrance': 'مدخل بنك',
        'emergency_exit': 'مخرج طارئ',
        'atm': 'ATM',
        'waiting_area': 'منطقة انتظار',
        'employees': 'موظفين',
        'right': 'يمين',
        'left': 'شمال',
        'center': 'وسط',
        'front': 'أمامي',
        'back': 'خلفي',
        'morning': 'صباحي',
        'evening': 'مسائي',
        '24hours': '24 ساعة',
        'active': 'نشط',
        'inactive': 'غير نشط',
        'maintenance': 'صيانة',
    },
    'en': {
        'app_title': 'Dashboard - Camera Management',
        'login_title': 'Login',
        'register_title': 'Create New Account',
        'username': 'Username',
        'password': 'Password',
        'confirm_password': 'Confirm Password',
        'login': 'Login',
        'register': 'Register',
        'back_to_login': 'Back to Login',
        'logout': 'Logout',
        'dashboard': 'Dashboard',
        'cameras': 'Cameras',
        'add_camera': 'Add Camera',
        'statistics': 'Statistics',
        'add': 'Add',
        'update': 'Update',
        'delete': 'Delete',
        'clear': 'Clear',
        'search': 'Search',
        'export_excel': 'Export Excel',
        'import_excel': 'Import Excel',
        'take_screenshot': 'Upload Image',
        'upload_image': 'Choose Image from Device',
        'camera_number': 'Camera Number',
        'branch_name': 'Branch Name',
        'branch_code': 'Branch Code',
        'device_ip': 'Device IP',
        'dvr_name': 'DVR Name',
        'username_field': 'Username',
        'password_field': 'Password',
        'location': 'Location Details',
        'location_type': 'Location Type',
        'direction': 'Direction',
        'working_hours': 'Working Hours',
        'model': 'Model',
        'serial_number': 'Serial Number',
        'resolution': 'Resolution',
        'features': 'Features',
        'select_features': 'Select Features',
        'night_vision': 'Night Vision',
        'motion_detection': 'Motion Detection',
        'audio': 'Audio',
        'ptz': 'PTZ (Pan-Tilt-Zoom)',
        'waterproof': 'Waterproof',
        'vandal_proof': 'Vandal Proof',
        'face_recognition': 'Face Recognition',
        'license_plate': 'License Plate Recognition',
        'wide_angle': 'Wide Angle',
        'zoom': 'Optical Zoom',
        'two_way_audio': 'Two-Way Audio',
        'smart_tracking': 'Smart Tracking',
        'lens_layout': 'Lens Type',
        'rtsp_url': 'RTSP URL',
        'install_date': 'Installation Date',
        'warranty_expiry': 'Warranty Expiry',
        'recording_capacity': 'Recording Capacity',
        'status': 'Status',
        'notes': 'Notes',
        'screenshot': 'Screenshot',
        'ready': 'Ready',
        'total_cameras': 'Total Cameras',
        'active_cameras': 'Active Cameras',
        'inactive_cameras': 'Inactive Cameras',
        'maintenance_cameras': 'Maintenance Cameras',
        'recent_cameras': 'Recent Cameras',
        'view_all': 'View All',
        'export_pdf': 'Export PDF',
        'print_report': 'Print Report',
        'filter': 'Filter',
        'all_status': 'All Status',
        'all_locations': 'All Locations',
        'apply_filter': 'Apply Filter',
        'reset_filter': 'Reset Filter',
        'camera_details': 'Camera Details',
        'technical_info': 'Technical Info',
        'system_info': 'System Info',
        'language': 'Language',
        'users_management': 'Users Management',
        'add_user': 'Add User',
        'edit_user': 'Edit User',
        'delete_user': 'Delete User',
        'role': 'Role',
        'admin': 'Admin',
        'moderator': 'Moderator',
        'user': 'User',
        'created_by': 'Created By',
        'created_at': 'Created At',
        'change_password': 'Change Password',
        'new_password': 'New Password',
        'user_status': 'User Status',
        'active_user': 'Active',
        'inactive_user': 'Inactive',
        'vault': 'Vault',
        'teller': 'Teller',
        'customer_service': 'Customer Service',
        'entrance': 'Bank Entrance',
        'emergency_exit': 'Emergency Exit',
        'atm': 'ATM',
        'waiting_area': 'Waiting Area',
        'employees': 'Employees',
        'right': 'Right',
        'left': 'Left',
        'center': 'Center',
        'front': 'Front',
        'back': 'Back',
        'morning': 'Morning',
        'evening': 'Evening',
        '24hours': '24 Hours',
        'active': 'Active',
        'inactive': 'Inactive',
        'maintenance': 'Maintenance',
    }
}

# ==================== USER MANAGEMENT ====================
class UserManager:
    def __init__(self):
        self.users_file = "users.json"
        self.users = self.load_users()
    
    def load_users(self):
        if os.path.exists(self.users_file):
            try:
                with open(self.users_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return {}
        return {}
    
    def save_users(self):
        with open(self.users_file, 'w', encoding='utf-8') as f:
            json.dump(self.users, f, ensure_ascii=False, indent=2)
    
    def hash_password(self, password):
        return hashlib.sha256(password.encode()).hexdigest()
    
    def register(self, username, password):
        if username in self.users:
            return False, "Username already exists"
        self.users[username] = {
            'password': self.hash_password(password),
            'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        self.save_users()
        return True, "Registration successful"
    
    def create_default_admin(self):
        if not self.users:
            self.users['admin'] = {
                'password': self.hash_password('admin'),
                'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            self.save_users()
            return True
        return False
    
    def login(self, username, password):
        if username not in self.users:
            return False, "User not found"
        if self.users[username]['password'] == self.hash_password(password):
            return True, "Login successful"
        return False, "Invalid password"

# ==================== LOGIN WINDOW ====================
class LoginWindow:
    def __init__(self, root, on_success):
        self.root = root
        self.on_success = on_success
        self.user_manager = UserManager()
        self.lang = 'ar'
        
        if self.user_manager.create_default_admin():
            print("Default admin account created: username='admin', password='admin'")
        
        self.root.title(self.t('login_title'))
        self.root.geometry("500x650")
        self.root.configure(bg=COLORS['bg_main'])
        self.root.resizable(False, False)
        
        self.center_window()
        self.setup_login_ui()
    
    def center_window(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def t(self, key):
        return TRANSLATIONS[self.lang].get(key, key)
    
    def clear_window(self):
        for widget in self.root.winfo_children():
            widget.destroy()
    
    def setup_login_ui(self):
        self.clear_window()
        
        main_frame = tk.Frame(self.root, bg=COLORS['bg_card'])
        main_frame.place(relx=0.5, rely=0.5, anchor='center', width=420, height=550)
        
        # Logo
        logo_frame = tk.Frame(main_frame, bg=COLORS['bg_card'], height=120)
        logo_frame.pack(fill='x')
        logo_frame.pack_propagate(False)
        
        icon_label = tk.Label(
            logo_frame,
            text="🎥",
            font=("Arial", 60),
            bg=COLORS['bg_card'],
            fg=COLORS['accent_blue']
        )
        icon_label.pack(pady=(25, 5))
        
        # Animate icon
        self.animate_icon(icon_label)
        
        title = tk.Label(
            logo_frame,
            text=self.t('app_title'),
            font=("Arial", 16, "bold"),
            bg=COLORS['bg_card'],
            fg=COLORS['text_primary']
        )
        title.pack(pady=(0, 10))
        
        # Form
        form_frame = tk.Frame(main_frame, bg=COLORS['bg_card'])
        form_frame.pack(pady=30, padx=40, fill='both', expand=True)
        
        subtitle = tk.Label(
            form_frame,
            text=self.t('login_title'),
            font=("Arial", 18, "bold"),
            bg=COLORS['bg_card'],
            fg=COLORS['text_primary']
        )
        subtitle.pack(pady=(0, 20))
        
        # Info
        info_frame = tk.Frame(form_frame, bg=COLORS['bg_main'])
        info_frame.pack(fill='x', pady=(0, 25), padx=5)
        
        info_text = "💡 الحساب الافتراضي:\nadmin / admin" if self.lang == 'ar' else "💡 Default Account:\nadmin / admin"
        info_label = tk.Label(
            info_frame,
            text=info_text,
            font=("Arial", 9),
            bg=COLORS['bg_main'],
            fg=COLORS['text_dim'],
            justify='center'
        )
        info_label.pack(pady=10)
        
        # Username
        username_label = tk.Label(
            form_frame,
            text=self.t('username'),
            font=("Arial", 12, "bold"),
            bg=COLORS['bg_card'],
            fg=COLORS['text_secondary'],
            anchor='w'
        )
        username_label.pack(fill='x', pady=(0, 5))
        
        self.username_entry = tk.Entry(
            form_frame,
            font=("Arial", 13),
            bg=COLORS['bg_input'],
            fg=COLORS['text_primary'],
            insertbackground=COLORS['accent_blue'],
            relief='flat',
            highlightthickness=1,
            highlightbackground=COLORS['border'],
            highlightcolor=COLORS['accent_blue']
        )
        self.username_entry.pack(fill='x', ipady=12, pady=(0, 20))
        
        # Add focus animation
        self.add_entry_animation(self.username_entry)
        
        # Password
        password_label = tk.Label(
            form_frame,
            text=self.t('password'),
            font=("Arial", 12, "bold"),
            bg=COLORS['bg_card'],
            fg=COLORS['text_secondary'],
            anchor='w'
        )
        password_label.pack(fill='x', pady=(0, 5))
        
        self.password_entry = tk.Entry(
            form_frame,
            font=("Arial", 13),
            show="●",
            bg=COLORS['bg_input'],
            fg=COLORS['text_primary'],
            insertbackground=COLORS['accent_blue'],
            relief='flat',
            highlightthickness=1,
            highlightbackground=COLORS['border'],
            highlightcolor=COLORS['accent_blue']
        )
        self.password_entry.pack(fill='x', ipady=12, pady=(0, 30))
        
        # Add focus animation
        self.add_entry_animation(self.password_entry)
        
        # Login button
        login_btn = tk.Button(
            form_frame,
            text="🔑 " + self.t('login'),
            command=self.login,
            bg=COLORS['accent_blue'],
            fg=COLORS['text_primary'],
            font=("Arial", 14, "bold"),
            relief='flat',
            cursor='hand2',
            activebackground=COLORS['accent_blue_dark'],
            activeforeground=COLORS['text_primary'],
            borderwidth=0
        )
        login_btn.pack(fill='x', ipady=14, pady=(0, 10))
        
        # Add hover animation
        self.add_button_hover(login_btn, COLORS['accent_blue'], COLORS['accent_blue_dark'])
        
        # Register button
        register_btn = tk.Button(
            form_frame,
            text="✨ " + self.t('register'),
            command=self.setup_register_ui,
            bg=COLORS['accent_green'],
            fg=COLORS['text_primary'],
            font=("Arial", 14, "bold"),
            relief='flat',
            cursor='hand2',
            activebackground='#246e44',
            activeforeground=COLORS['text_primary'],
            borderwidth=0
        )
        register_btn.pack(fill='x', ipady=14, pady=(0, 20))
        
        # Add hover animation
        self.add_button_hover(register_btn, COLORS['accent_green'], '#246e44')
        
        # Language
        lang_btn = tk.Button(
            form_frame,
            text="🌐 العربية | English",
            command=self.toggle_language,
            bg=COLORS['bg_main'],
            fg=COLORS['text_dim'],
            font=("Arial", 9),
            relief='flat',
            cursor='hand2',
            padx=15,
            pady=5,
            borderwidth=0
        )
        lang_btn.pack()
        
        self.password_entry.bind('<Return>', lambda e: self.login())
        self.username_entry.focus()
    
    def animate_icon(self, label):
        """Animate icon with pulsing effect"""
        def pulse():
            sizes = [60, 63, 66, 63, 60]
            for size in sizes:
                try:
                    label.config(font=("Arial", size))
                    label.update()
                    time.sleep(0.1)
                except:
                    break
        
        def start_animation():
            while True:
                pulse()
                time.sleep(2)
        
        thread = threading.Thread(target=start_animation, daemon=True)
        thread.start()
    
    def add_entry_animation(self, entry):
        """Add focus animation to entry"""
        original_bg = entry.cget('bg')
        
        def on_focus_in(e):
            entry.config(highlightthickness=2)
        
        def on_focus_out(e):
            entry.config(highlightthickness=1)
        
        entry.bind('<FocusIn>', on_focus_in)
        entry.bind('<FocusOut>', on_focus_out)
    
    def add_button_hover(self, button, normal_color, hover_color):
        """Add hover effect to button"""
        def on_enter(e):
            button.config(bg=hover_color)
        
        def on_leave(e):
            button.config(bg=normal_color)
        
        button.bind('<Enter>', on_enter)
        button.bind('<Leave>', on_leave)
    
    def setup_register_ui(self):
        self.clear_window()
        
        main_frame = tk.Frame(self.root, bg=COLORS['bg_card'])
        main_frame.place(relx=0.5, rely=0.5, anchor='center', width=420, height=520)
        
        # Header
        header_frame = tk.Frame(main_frame, bg=COLORS['accent_green'], height=80)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)
        
        title = tk.Label(
            header_frame,
            text="✨ " + self.t('register_title'),
            font=("Arial", 16, "bold"),
            bg=COLORS['accent_green'],
            fg=COLORS['text_primary']
        )
        title.pack(pady=25)
        
        # Form
        form_frame = tk.Frame(main_frame, bg=COLORS['bg_card'])
        form_frame.pack(pady=30, padx=40, fill='both', expand=True)
        
        # Username
        username_label = tk.Label(
            form_frame,
            text=self.t('username'),
            font=("Arial", 10, "bold"),
            bg=COLORS['bg_card'],
            fg=COLORS['text_secondary'],
            anchor='w'
        )
        username_label.pack(fill='x', pady=(0, 5))
        
        self.reg_username_entry = tk.Entry(
            form_frame,
            font=("Arial", 11),
            bg=COLORS['bg_input'],
            fg=COLORS['text_primary'],
            insertbackground=COLORS['accent_green'],
            relief='flat',
            highlightthickness=1,
            highlightbackground=COLORS['border'],
            highlightcolor=COLORS['accent_green']
        )
        self.reg_username_entry.pack(fill='x', ipady=10, pady=(0, 15))
        
        # Password
        password_label = tk.Label(
            form_frame,
            text=self.t('password'),
            font=("Arial", 10, "bold"),
            bg=COLORS['bg_card'],
            fg=COLORS['text_secondary'],
            anchor='w'
        )
        password_label.pack(fill='x', pady=(0, 5))
        
        self.reg_password_entry = tk.Entry(
            form_frame,
            font=("Arial", 11),
            show="●",
            bg=COLORS['bg_input'],
            fg=COLORS['text_primary'],
            insertbackground=COLORS['accent_green'],
            relief='flat',
            highlightthickness=1,
            highlightbackground=COLORS['border'],
            highlightcolor=COLORS['accent_green']
        )
        self.reg_password_entry.pack(fill='x', ipady=10, pady=(0, 15))
        
        # Confirm
        confirm_label = tk.Label(
            form_frame,
            text=self.t('confirm_password'),
            font=("Arial", 10, "bold"),
            bg=COLORS['bg_card'],
            fg=COLORS['text_secondary'],
            anchor='w'
        )
        confirm_label.pack(fill='x', pady=(0, 5))
        
        self.reg_confirm_entry = tk.Entry(
            form_frame,
            font=("Arial", 11),
            show="●",
            bg=COLORS['bg_input'],
            fg=COLORS['text_primary'],
            insertbackground=COLORS['accent_green'],
            relief='flat',
            highlightthickness=1,
            highlightbackground=COLORS['border'],
            highlightcolor=COLORS['accent_green']
        )
        self.reg_confirm_entry.pack(fill='x', ipady=10, pady=(0, 25))
        
        # Register
        register_btn = tk.Button(
            form_frame,
            text="✅ " + self.t('register'),
            command=self.register,
            bg=COLORS['accent_green'],
            fg=COLORS['text_primary'],
            font=("Arial", 12, "bold"),
            relief='flat',
            cursor='hand2',
            activebackground='#246e44',
            borderwidth=0
        )
        register_btn.pack(fill='x', ipady=12, pady=(0, 10))
        
        # Back
        back_btn = tk.Button(
            form_frame,
            text="← " + self.t('back_to_login'),
            command=self.setup_login_ui,
            bg=COLORS['border'],
            fg=COLORS['text_primary'],
            font=("Arial", 11, "bold"),
            relief='flat',
            cursor='hand2',
            borderwidth=0
        )
        back_btn.pack(fill='x', ipady=10)
        
        self.reg_username_entry.focus()
    
    def toggle_language(self):
        self.lang = 'en' if self.lang == 'ar' else 'ar'
        self.setup_login_ui()
    
    def login(self):
        username = self.username_entry.get().strip()
        password = self.password_entry.get()
        
        if not username or not password:
            messagebox.showwarning(
                "تحذير" if self.lang == 'ar' else "Warning",
                "يرجى ملء جميع الحقول" if self.lang == 'ar' else "Please fill all fields"
            )
            return
        
        success, message, role = self.user_manager.login(username, password)
        if success:
            self.root.withdraw()
            self.on_success(username, self.lang, role)
        else:
            messagebox.showerror(
                "خطأ" if self.lang == 'ar' else "Error",
                "اسم المستخدم أو كلمة المرور غير صحيحة" if self.lang == 'ar' 
                else "Invalid username or password"
            )
    
    def register(self):
        username = self.reg_username_entry.get().strip()
        password = self.reg_password_entry.get()
        confirm = self.reg_confirm_entry.get()
        
        if not username or not password or not confirm:
            messagebox.showwarning(
                "تحذير" if self.lang == 'ar' else "Warning",
                "يرجى ملء جميع الحقول" if self.lang == 'ar' else "Please fill all fields"
            )
            return
        
        if len(password) < 4:
            messagebox.showwarning(
                "تحذير" if self.lang == 'ar' else "Warning",
                "كلمة المرور يجب أن تكون 4 أحرف على الأقل" if self.lang == 'ar' 
                else "Password must be at least 4 characters"
            )
            return
        
        if password != confirm:
            messagebox.showerror(
                "خطأ" if self.lang == 'ar' else "Error",
                "كلمات المرور غير متطابقة" if self.lang == 'ar' 
                else "Passwords do not match"
            )
            return
        
        success, message = self.user_manager.register(username, password)
        if success:
            messagebox.showinfo(
                "نجح" if self.lang == 'ar' else "Success",
                "تم إنشاء الحساب بنجاح!" if self.lang == 'ar' 
                else "Account created successfully!"
            )
            self.setup_login_ui()
        else:
            messagebox.showerror(
                "خطأ" if self.lang == 'ar' else "Error",
                "اسم المستخدم موجود بالفعل" if self.lang == 'ar' 
                else "Username already exists"
            )

# ==================== DASHBOARD ====================
class DashboardApp:
    def __init__(self, root, username, lang='ar', role='user'):
        self.root = root
        self.username = username
        self.lang = lang
        self.role = role
        self.current_page = 'dashboard'
        self.user_manager = UserManager()
        
        self.root.title(self.t('app_title'))
        self.root.geometry("1600x900")
        try:
            self.root.state('zoomed')
        except:
            pass
        self.root.configure(bg=COLORS['bg_main'])
        
        self.cameras_data = []
        self.data_file = f"{username}_cameras.json"
        self.load_data()
        
        self.current_screenshot = None
        self.selected_features = []
        self.setup_ui()
    
    def t(self, key):
        return TRANSLATIONS[self.lang].get(key, key)
    
    def setup_ui(self):
        # Sidebar
        sidebar = tk.Frame(self.root, bg=COLORS['bg_sidebar'], width=250)
        sidebar.pack(side='left', fill='y')
        sidebar.pack_propagate(False)
        
        # Logo
        logo_frame = tk.Frame(sidebar, bg=COLORS['bg_sidebar'], height=100)
        logo_frame.pack(fill='x')
        logo_frame.pack_propagate(False)
        
        tk.Label(
            logo_frame,
            text="🎥",
            font=("Arial", 40),
            bg=COLORS['bg_sidebar'],
            fg=COLORS['accent_blue']
        ).pack(pady=(20, 0))
        
        tk.Label(
            logo_frame,
            text=self.t('app_title'),
            font=("Arial", 13, "bold"),
            bg=COLORS['bg_sidebar'],
            fg=COLORS['text_secondary']
        ).pack(pady=(5, 15))
        
        # Menu
        menu_items = [
            ('dashboard', '📊', self.t('dashboard')),
            ('cameras', '🎥', self.t('cameras')),
            ('add_camera', '➕', self.t('add_camera')),
            ('statistics', '📈', self.t('statistics'))
        ]
        
        # Add users management for admin only
        if self.role == 'admin':
            menu_items.append(('users', '👥', self.t('users_management')))
        
        self.menu_buttons = {}
        for page_id, icon, text in menu_items:
            btn = tk.Button(
                sidebar,
                text=f"{icon}  {text}",
                command=lambda p=page_id: self.switch_page(p),
                bg=COLORS['bg_sidebar'],
                fg=COLORS['text_dim'],
                font=("Arial", 13, "bold"),
                relief='flat',
                cursor='hand2',
                anchor='w',
                padx=25,
                pady=18,
                borderwidth=0,
                activebackground=COLORS['hover'],
                activeforeground=COLORS['text_primary']
            )
            btn.pack(fill='x', pady=2)
            self.menu_buttons[page_id] = btn
            
            # Add hover effect
            self.add_menu_hover(btn)
        
        # Spacer
        tk.Frame(sidebar, bg=COLORS['bg_sidebar']).pack(fill='both', expand=True)
        
        # User section
        user_frame = tk.Frame(sidebar, bg=COLORS['bg_main'])
        user_frame.pack(fill='x', pady=20, padx=15)
        
        tk.Label(
            user_frame,
            text=f"👤 {self.username} ({self.t(self.role)})",
            font=("Arial", 11, "bold"),
            bg=COLORS['bg_main'],
            fg=COLORS['text_secondary']
        ).pack(pady=10)
        
        # Language
        lang_btn = tk.Button(
            user_frame,
            text="🌐",
            command=self.toggle_language,
            bg=COLORS['accent_blue'],
            fg=COLORS['text_primary'],
            font=("Arial", 12, "bold"),
            relief='flat',
            cursor='hand2',
            width=4,
            pady=8,
            borderwidth=0
        )
        lang_btn.pack(side='left', padx=5, expand=True)
        self.add_button_animation(lang_btn, COLORS['accent_blue'], COLORS['accent_blue_dark'])
        
        # Logout
        logout_btn = tk.Button(
            user_frame,
            text="🚪",
            command=self.logout,
            bg=COLORS['accent_red'],
            fg=COLORS['text_primary'],
            font=("Arial", 12, "bold"),
            relief='flat',
            cursor='hand2',
            width=4,
            pady=8,
            borderwidth=0
        )
        logout_btn.pack(side='right', padx=5, expand=True)
        self.add_button_animation(logout_btn, COLORS['accent_red'], '#8b1c1c')
        
        # Main content
        self.content_frame = tk.Frame(self.root, bg=COLORS['bg_main'])
        self.content_frame.pack(side='right', fill='both', expand=True)
        
        self.switch_page('dashboard')
    
    def add_menu_hover(self, button):
        """Add hover effect to menu button"""
        original_bg = button.cget('bg')
        
        def on_enter(e):
            if button.cget('bg') != COLORS['accent_blue']:
                button.config(bg=COLORS['hover'])
        
        def on_leave(e):
            if button.cget('bg') != COLORS['accent_blue']:
                button.config(bg=original_bg)
        
        button.bind('<Enter>', on_enter)
        button.bind('<Leave>', on_leave)
    
    def add_button_animation(self, button, normal_color, hover_color):
        """Add hover animation to button with scale effect"""
        def on_enter(e):
            button.config(bg=hover_color)
            # Add slight scale effect through padding
            current_pady = button.cget('pady')
            button.config(pady=int(current_pady) + 2)
        
        def on_leave(e):
            button.config(bg=normal_color)
            button.config(pady=8)
        
        button.bind('<Enter>', on_enter)
        button.bind('<Leave>', on_leave)
    
    def animate_card(self, card):
        """Animate card appearance"""
        card.config(relief='flat')
        
        def animate():
            for i in range(5):
                card.config(highlightthickness=i % 2)
                card.update()
                time.sleep(0.05)
        
        threading.Thread(target=animate, daemon=True).start()
    
    def switch_page(self, page):
        self.current_page = page
        
        # Update menu buttons
        for btn_id, btn in self.menu_buttons.items():
            if btn_id == page:
                btn.config(bg=COLORS['accent_blue'], fg=COLORS['text_primary'])
            else:
                btn.config(bg=COLORS['bg_sidebar'], fg=COLORS['text_dim'])
        
        # Clear content
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        # Load page
        if page == 'dashboard':
            self.show_dashboard()
        elif page == 'cameras':
            self.show_cameras()
        elif page == 'add_camera':
            self.show_add_camera()
        elif page == 'statistics':
            self.show_statistics()
        elif page == 'users' and self.role == 'admin':
            self.show_users_management()
    
    def show_dashboard(self):
        # Header
        header = tk.Frame(self.content_frame, bg=COLORS['bg_main'])
        header.pack(fill='x', padx=30, pady=20)
        
        tk.Label(
            header,
            text="📊 " + self.t('dashboard'),
            font=("Arial", 28, "bold"),
            bg=COLORS['bg_main'],
            fg=COLORS['text_primary']
        ).pack(side='left', anchor='w')
        
        # Export button
        tk.Button(
            header,
            text="📊 " + self.t('export_excel'),
            command=self.export_to_excel,
            bg=COLORS['accent_green'],
            fg=COLORS['text_primary'],
            font=("Arial", 11, "bold"),
            relief='flat',
            cursor='hand2',
            padx=20,
            pady=10,
            borderwidth=0
        ).pack(side='right', padx=5)
        
        tk.Button(
            header,
            text="🖨️ " + self.t('print_report'),
            command=self.print_report,
            bg=COLORS['accent_blue'],
            fg=COLORS['text_primary'],
            font=("Arial", 11, "bold"),
            relief='flat',
            cursor='hand2',
            padx=20,
            pady=10,
            borderwidth=0
        ).pack(side='right', padx=5)
        
        # Stats cards
        stats_frame = tk.Frame(self.content_frame, bg=COLORS['bg_main'])
        stats_frame.pack(fill='x', padx=30, pady=10)
        
        total = len(self.cameras_data)
        active = sum(1 for c in self.cameras_data if c.get('status') == 'active')
        inactive = sum(1 for c in self.cameras_data if c.get('status') == 'inactive')
        maintenance = sum(1 for c in self.cameras_data if c.get('status') == 'maintenance')
        
        stats = [
            (self.t('total_cameras'), str(total), '🎥', COLORS['accent_blue']),
            (self.t('active_cameras'), str(active), '✅', COLORS['accent_green']),
            (self.t('inactive_cameras'), str(inactive), '⭕', COLORS['text_dim']),
            (self.t('maintenance_cameras'), str(maintenance), '🔧', COLORS['accent_red'])
        ]
        
        for title, value, icon, color in stats:
            card = tk.Frame(stats_frame, bg=COLORS['bg_card'], relief='flat', bd=1, highlightbackground=COLORS['border'], highlightthickness=1)
            card.pack(side='left', fill='both', expand=True, padx=8)
            
            # Animate card on load
            self.animate_card(card)
            
            icon_label = tk.Label(
                card,
                text=icon,
                font=("Arial", 40),
                bg=COLORS['bg_card'],
                fg=color
            )
            icon_label.pack(pady=(20, 5))
            
            # Animate icon rotation
            self.add_icon_pulse(icon_label)
            
            tk.Label(
                card,
                text=value,
                font=("Arial", 32, "bold"),
                bg=COLORS['bg_card'],
                fg=COLORS['text_primary']
            ).pack()
            
            tk.Label(
                card,
                text=title,
                font=("Arial", 12),
                bg=COLORS['bg_card'],
                fg=COLORS['text_dim']
            ).pack(pady=(5, 20))
            
            # Add hover effect
            self.add_card_hover(card)
        
        # Recent cameras
        recent_frame = tk.Frame(self.content_frame, bg=COLORS['bg_card'])
        recent_frame.pack(fill='both', expand=True, padx=30, pady=20)
        
        recent_header = tk.Frame(recent_frame, bg=COLORS['bg_card'])
        recent_header.pack(fill='x', padx=20, pady=15)
        
        tk.Label(
            recent_header,
            text="📋 " + self.t('recent_cameras'),
            font=("Arial", 18, "bold"),
            bg=COLORS['bg_card'],
            fg=COLORS['text_primary']
        ).pack(side='left', anchor='w')
        
        view_all_btn = tk.Button(
            recent_header,
            text=self.t('view_all') + " →",
            command=lambda: self.switch_page('cameras'),
            bg=COLORS['bg_input'],
            fg=COLORS['accent_blue'],
            font=("Arial", 11, "bold"),
            relief='flat',
            cursor='hand2',
            padx=18,
            pady=10,
            borderwidth=0
        )
        view_all_btn.pack(side='right')
        self.add_button_animation(view_all_btn, COLORS['bg_input'], COLORS['hover'])
        
        # Table
        table_frame = tk.Frame(recent_frame, bg=COLORS['bg_card'])
        table_frame.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        style = ttk.Style()
        style.theme_use('clam')
        style.configure(
            "Dashboard.Treeview",
            background=COLORS['bg_main'],
            foreground=COLORS['text_secondary'],
            fieldbackground=COLORS['bg_main'],
            borderwidth=0,
            font=("Arial", 11),
            rowheight=30
        )
        style.configure(
            "Dashboard.Treeview.Heading",
            background=COLORS['bg_sidebar'],
            foreground=COLORS['text_primary'],
            borderwidth=0,
            relief="flat",
            font=("Arial", 12, "bold")
        )
        style.map(
            "Dashboard.Treeview",
            background=[('selected', COLORS['accent_blue'])],
            foreground=[('selected', COLORS['text_primary'])]
        )
        
        columns = ("camera_number", "branch_name", "device_ip", "status")
        tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show='headings',
            height=10,
            style="Dashboard.Treeview"
        )
        
        for col in columns:
            tree.heading(col, text=self.t(col), anchor='center')
            tree.column(col, width=200, anchor='center')
        
        for camera in self.cameras_data[-10:]:
            values = (
                camera.get('camera_number', ''),
                camera.get('branch_name', ''),
                camera.get('device_ip', ''),
                self.t(camera.get('status', ''))
            )
            tree.insert('', 'end', values=values)
        
        tree.pack(fill='both', expand=True)
        
        # Bind double click to view details
        tree.bind('<Double-1>', lambda e: self.show_camera_details(tree))
    
    def add_icon_pulse(self, label):
        """Add pulsing animation to icon"""
        def pulse():
            original_size = 40
            while True:
                for size in [40, 43, 40]:
                    try:
                        label.config(font=("Arial", size))
                        label.update()
                        time.sleep(0.3)
                    except:
                        return
                time.sleep(2)
        
        threading.Thread(target=pulse, daemon=True).start()
    
    def add_card_hover(self, card):
        """Add hover effect to card"""
        original_bg = card.cget('bg')
        
        def on_enter(e):
            card.config(highlightthickness=2, highlightbackground=COLORS['accent_blue'])
        
        def on_leave(e):
            card.config(highlightthickness=1, highlightbackground=COLORS['border'])
        
        card.bind('<Enter>', on_enter)
        card.bind('<Leave>', on_leave)
        
        # Bind to all children
        for child in card.winfo_children():
            child.bind('<Enter>', on_enter)
            child.bind('<Leave>', on_leave)
    
    def show_cameras(self):
        # Header
        header = tk.Frame(self.content_frame, bg=COLORS['bg_main'])
        header.pack(fill='x', padx=30, pady=20)
        
        tk.Label(
            header,
            text="🎥 " + self.t('cameras'),
            font=("Arial", 28, "bold"),
            bg=COLORS['bg_main'],
            fg=COLORS['text_primary']
        ).pack(side='left')
        
        # Filter button
        filter_btn = tk.Button(
            header,
            text="🔍 " + self.t('filter'),
            command=self.show_filter_dialog,
            bg=COLORS['accent_blue'],
            fg=COLORS['text_primary'],
            font=("Arial", 12, "bold"),
            relief='flat',
            cursor='hand2',
            padx=18,
            pady=10,
            borderwidth=0
        )
        filter_btn.pack(side='right', padx=5)
        self.add_button_animation(filter_btn, COLORS['accent_blue'], COLORS['accent_blue_dark'])
        
        # Search and actions
        actions_frame = tk.Frame(header, bg=COLORS['bg_main'])
        actions_frame.pack(side='right')
        
        # Search
        search_frame = tk.Frame(actions_frame, bg=COLORS['bg_input'], highlightthickness=1, highlightbackground=COLORS['border'])
        search_frame.pack(side='left', padx=5)
        
        tk.Label(
            search_frame,
            text="🔍",
            bg=COLORS['bg_input'],
            fg=COLORS['text_dim']
        ).pack(side='left', padx=8)
        
        self.search_var = tk.StringVar()
        self.search_var.trace('w', lambda *args: self.search_cameras())
        search_entry = tk.Entry(
            search_frame,
            textvariable=self.search_var,
            font=("Arial", 11),
            width=25,
            bg=COLORS['bg_input'],
            fg=COLORS['text_primary'],
            relief='flat',
            insertbackground=COLORS['accent_blue'],
            borderwidth=0
        )
        search_entry.pack(side='left', ipady=8, padx=(0, 8))
        
        # Buttons
        btn_frame = tk.Frame(actions_frame, bg=COLORS['bg_main'])
        btn_frame.pack(side='left', padx=5)
        
        action_btns = [
            ("📊", self.export_to_excel, COLORS['accent_green']),
            ("📥", self.import_from_excel, COLORS['accent_blue']),
            ("❌", self.delete_camera, COLORS['accent_red'])
        ]
        
        for icon, cmd, color in action_btns:
            action_btn = tk.Button(
                btn_frame,
                text=icon,
                command=cmd,
                bg=color,
                fg=COLORS['text_primary'],
                font=("Arial", 14, "bold"),
                relief='flat',
                cursor='hand2',
                width=3,
                pady=10,
                borderwidth=0
            )
            action_btn.pack(side='left', padx=2)
            
            # Add hover animation
            hover_colors = {
                COLORS['accent_green']: '#1d7a47',
                COLORS['accent_blue']: COLORS['accent_blue_dark'],
                COLORS['accent_red']: '#8b1c1c'
            }
            self.add_button_animation(action_btn, color, hover_colors.get(color, color))
        
        # Table
        table_container = tk.Frame(self.content_frame, bg=COLORS['bg_card'])
        table_container.pack(fill='both', expand=True, padx=30, pady=(0, 20))
        
        table_frame = tk.Frame(table_container, bg=COLORS['bg_card'])
        table_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        v_scrollbar = ttk.Scrollbar(table_frame, orient='vertical')
        h_scrollbar = ttk.Scrollbar(table_frame, orient='horizontal')
        
        columns = (
            "camera_number", "branch_name", "branch_code", "device_ip",
            "location", "location_type", "direction", "status"
        )
        
        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show='headings',
            yscrollcommand=v_scrollbar.set,
            xscrollcommand=h_scrollbar.set,
            height=20,
            style="Dashboard.Treeview"
        )
        
        v_scrollbar.config(command=self.tree.yview)
        h_scrollbar.config(command=self.tree.xview)
        
        for col in columns:
            self.tree.heading(col, text=self.t(col), anchor='center')
            self.tree.column(col, width=130, anchor='center')
        
        self.tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        self.tree.bind('<Double-1>', self.on_row_double_click)
        
        # Bind right-click for context menu
        self.tree.bind('<Button-3>', self.show_context_menu)
        
        self.refresh_table()
    
    def show_add_camera(self):
        # Header
        header = tk.Frame(self.content_frame, bg=COLORS['bg_main'])
        header.pack(fill='x', padx=30, pady=20)
        
        tk.Label(
            header,
            text="➕ " + self.t('add_camera'),
            font=("Arial", 24, "bold"),
            bg=COLORS['bg_main'],
            fg=COLORS['text_primary']
        ).pack(side='left', anchor='w')
        
        # Export button
        tk.Button(
            header,
            text="📊 " + self.t('export_excel'),
            command=self.export_to_excel,
            bg=COLORS['accent_green'],
            fg=COLORS['text_primary'],
            font=("Arial", 11, "bold"),
            relief='flat',
            cursor='hand2',
            padx=20,
            pady=10,
            borderwidth=0
        ).pack(side='right')
        
        # Form container
        form_container = tk.Frame(self.content_frame, bg=COLORS['bg_main'])
        form_container.pack(fill='both', expand=True, padx=30, pady=(0, 20))
        
        # Left: Form
        left_frame = tk.Frame(form_container, bg=COLORS['bg_card'])
        left_frame.pack(side='left', fill='both', expand=True, padx=(0, 15))
        
        # Scrollable form
        form_canvas = tk.Canvas(left_frame, bg=COLORS['bg_card'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(left_frame, orient="vertical", command=form_canvas.yview)
        form_frame = tk.Frame(form_canvas, bg=COLORS['bg_card'])
        
        form_frame.bind(
            "<Configure>",
            lambda e: form_canvas.configure(scrollregion=form_canvas.bbox("all"))
        )
        
        form_canvas.create_window((0, 0), window=form_frame, anchor="nw")
        form_canvas.configure(yscrollcommand=scrollbar.set)
        
        form_canvas.pack(side="left", fill="both", expand=True, padx=20, pady=20)
        scrollbar.pack(side="right", fill="y")
        
        def _on_mousewheel(event):
            form_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        form_canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Form fields
        self.entries = {}
        
        fields = [
            ('camera_number', None),
            ('branch_name', None),
            ('branch_code', None),
            ('device_ip', None),
            ('dvr_name', None),
            ('username_field', None),
            ('password_field', None),
            ('location', None),
            ('location_type', ['vault', 'teller', 'customer_service', 'entrance', 
                              'emergency_exit', 'atm', 'waiting_area', 'employees']),
            ('direction', ['right', 'left', 'center', 'front', 'back']),
            ('working_hours', ['morning', 'evening', '24hours']),
            ('model', None),
            ('serial_number', None),
            ('features', ['night_vision', 'motion_detection', 'audio', 'ptz', 
                         'waterproof', 'vandal_proof', 'face_recognition', 
                         'license_plate', 'wide_angle', 'zoom', 'two_way_audio', 'smart_tracking']),
            ('lens_layout', None),
            ('rtsp_url', None),
            ('status', ['active', 'inactive', 'maintenance']),
            ('notes', None)
        ]
        
        for field_key, options in fields:
            field_frame = tk.Frame(form_frame, bg=COLORS['bg_card'])
            field_frame.pack(fill='x', pady=8)
            
            label = tk.Label(
                field_frame,
                text=self.t(field_key),
                font=("Arial", 10, "bold"),
                bg=COLORS['bg_card'],
                fg=COLORS['text_secondary'],
                anchor='w',
                width=18
            )
            label.pack(side='left', padx=5)
            
            # Special handling for features (multi-select)
            if field_key == 'features':
                # Create button to open features selection
                features_btn_frame = tk.Frame(field_frame, bg=COLORS['bg_input'], 
                                             highlightthickness=1, 
                                             highlightbackground=COLORS['border'])
                features_btn_frame.pack(side='left', fill='x', expand=True, padx=5)
                
                self.selected_features = []
                
                features_display = tk.Label(
                    features_btn_frame,
                    text=self.t('select_features'),
                    font=("Arial", 9),
                    bg=COLORS['bg_input'],
                    fg=COLORS['text_dim'],
                    anchor='w',
                    width=28
                )
                features_display.pack(side='left', padx=5, ipady=6)
                
                select_btn = tk.Button(
                    features_btn_frame,
                    text="...",
                    command=lambda: self.open_features_dialog(features_display),
                    bg=COLORS['accent_blue'],
                    fg=COLORS['text_primary'],
                    font=("Arial", 9, "bold"),
                    relief='flat',
                    cursor='hand2',
                    width=3,
                    borderwidth=0
                )
                select_btn.pack(side='right', padx=2, pady=2)
                
                self.entries[field_key] = features_display
                
            elif options:
                entry = ttk.Combobox(
                    field_frame,
                    font=("Arial", 10),
                    width=30,
                    state='readonly'
                )
                entry['values'] = [self.t(opt) for opt in options]
                if field_key == 'status':
                    entry.current(0)
                entry.pack(side='left', ipady=6, padx=5)
                self.entries[field_key] = entry
            else:
                entry = tk.Entry(
                    field_frame,
                    font=("Arial", 10),
                    width=32,
                    bg=COLORS['bg_input'],
                    fg=COLORS['text_primary'],
                    relief='flat',
                    highlightthickness=1,
                    highlightbackground=COLORS['border'],
                    highlightcolor=COLORS['accent_blue'],
                    insertbackground=COLORS['accent_blue']
                )
                entry.pack(side='left', ipady=6, padx=5)
                self.entries[field_key] = entry
        
        # Right: Screenshot & Actions
        right_frame = tk.Frame(form_container, bg=COLORS['bg_card'], width=350)
        right_frame.pack(side='right', fill='y')
        right_frame.pack_propagate(False)
        
        tk.Label(
            right_frame,
            text=self.t('screenshot'),
            font=("Arial", 14, "bold"),
            bg=COLORS['bg_card'],
            fg=COLORS['text_primary']
        ).pack(pady=(20, 15))
        
        self.screenshot_label = tk.Label(
            right_frame,
            text="🖼️\n" + self.t('upload_image'),
            bg=COLORS['bg_input'],
            fg=COLORS['text_dim'],
            relief='flat',
            width=30,
            height=10,
            font=("Arial", 10),
            highlightthickness=1,
            highlightbackground=COLORS['border']
        )
        self.screenshot_label.pack(pady=10, padx=20)
        
        tk.Button(
            right_frame,
            text="📁 " + self.t('take_screenshot'),
            command=self.upload_image,
            bg=COLORS['accent_blue'],
            fg=COLORS['text_primary'],
            font=("Arial", 11, "bold"),
            relief='flat',
            cursor='hand2',
            pady=10,
            borderwidth=0
        ).pack(pady=15, padx=20, fill='x')
        
        # Action buttons
        tk.Frame(right_frame, bg=COLORS['bg_card'], height=20).pack()
        
        buttons = [
            (self.t('add'), self.add_camera, COLORS['accent_green'], '➕'),
            (self.t('update'), self.update_camera, COLORS['accent_blue'], '✏️'),
            (self.t('delete'), self.delete_camera, COLORS['accent_red'], '❌'),
            (self.t('clear'), self.clear_form, COLORS['border'], '🗑️')
        ]
        
        for text, cmd, color, icon in buttons:
            tk.Button(
                right_frame,
                text=f"{icon} {text}",
                command=cmd,
                bg=color,
                fg=COLORS['text_primary'],
                font=("Arial", 11, "bold"),
                relief='flat',
                cursor='hand2',
                pady=12,
                borderwidth=0
            ).pack(pady=5, padx=20, fill='x')
    
    def show_statistics(self):
        # Header
        header = tk.Frame(self.content_frame, bg=COLORS['bg_main'])
        header.pack(fill='x', padx=30, pady=20)
        
        tk.Label(
            header,
            text="📈 " + self.t('statistics'),
            font=("Arial", 24, "bold"),
            bg=COLORS['bg_main'],
            fg=COLORS['text_primary']
        ).pack(side='left', anchor='w')
        
        # Export button
        tk.Button(
            header,
            text="📊 " + self.t('export_excel'),
            command=self.export_to_excel,
            bg=COLORS['accent_green'],
            fg=COLORS['text_primary'],
            font=("Arial", 11, "bold"),
            relief='flat',
            cursor='hand2',
            padx=20,
            pady=10,
            borderwidth=0
        ).pack(side='right')
        
        # Stats container
        stats_container = tk.Frame(self.content_frame, bg=COLORS['bg_main'])
        stats_container.pack(fill='both', expand=True, padx=30, pady=(0, 20))
        
        # Status distribution
        status_card = tk.Frame(stats_container, bg=COLORS['bg_card'])
        status_card.pack(fill='both', expand=True, pady=10)
        
        tk.Label(
            status_card,
            text="📊 " + ("توزيع الحالات" if self.lang == 'ar' else "Status Distribution"),
            font=("Arial", 16, "bold"),
            bg=COLORS['bg_card'],
            fg=COLORS['text_primary']
        ).pack(anchor='w', padx=20, pady=15)
        
        status_stats = tk.Frame(status_card, bg=COLORS['bg_card'])
        status_stats.pack(fill='x', padx=20, pady=(0, 20))
        
        active = sum(1 for c in self.cameras_data if c.get('status') == 'active')
        inactive = sum(1 for c in self.cameras_data if c.get('status') == 'inactive')
        maintenance = sum(1 for c in self.cameras_data if c.get('status') == 'maintenance')
        total = len(self.cameras_data) or 1
        
        stats = [
            (self.t('active'), active, f"{(active/total)*100:.1f}%", COLORS['accent_green']),
            (self.t('inactive'), inactive, f"{(inactive/total)*100:.1f}%", COLORS['text_dim']),
            (self.t('maintenance'), maintenance, f"{(maintenance/total)*100:.1f}%", COLORS['accent_red'])
        ]
        
        for label, count, percent, color in stats:
            stat_frame = tk.Frame(status_stats, bg=COLORS['bg_main'])
            stat_frame.pack(fill='x', pady=8)
            
            tk.Label(
                stat_frame,
                text=label,
                font=("Arial", 12, "bold"),
                bg=COLORS['bg_main'],
                fg=COLORS['text_secondary'],
                width=15,
                anchor='w'
            ).pack(side='left', padx=10)
            
            progress_bg = tk.Frame(stat_frame, bg=COLORS['bg_sidebar'], height=25)
            progress_bg.pack(side='left', fill='x', expand=True, padx=10)
            
            if total > 0:
                progress_width = int((count/total) * 400)
                tk.Frame(
                    progress_bg,
                    bg=color,
                    width=progress_width,
                    height=25
                ).place(x=0, y=0)
            
            tk.Label(
                stat_frame,
                text=f"{count} ({percent})",
                font=("Arial", 11, "bold"),
                bg=COLORS['bg_main'],
                fg=color,
                width=15
            ).pack(side='left', padx=10)
        
        # Location distribution
        location_card = tk.Frame(stats_container, bg=COLORS['bg_card'])
        location_card.pack(fill='both', expand=True, pady=10)
        
        tk.Label(
            location_card,
            text="📍 " + ("توزيع المواقع" if self.lang == 'ar' else "Location Distribution"),
            font=("Arial", 16, "bold"),
            bg=COLORS['bg_card'],
            fg=COLORS['text_primary']
        ).pack(anchor='w', padx=20, pady=15)
        
        location_stats = tk.Frame(location_card, bg=COLORS['bg_card'])
        location_stats.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        from collections import Counter
        locations = Counter(c.get('location_type', 'unknown') for c in self.cameras_data)
        
        for loc_type, count in locations.most_common():
            loc_frame = tk.Frame(location_stats, bg=COLORS['bg_main'], height=40)
            loc_frame.pack(fill='x', pady=5)
            loc_frame.pack_propagate(False)
            
            tk.Label(
                loc_frame,
                text=self.t(loc_type),
                font=("Arial", 11),
                bg=COLORS['bg_main'],
                fg=COLORS['text_secondary']
            ).pack(side='left', padx=15)
            
            tk.Label(
                loc_frame,
                text=str(count),
                font=("Arial", 13, "bold"),
                bg=COLORS['bg_main'],
                fg=COLORS['accent_blue']
            ).pack(side='right', padx=15)
    
    def toggle_language(self):
        self.lang = 'en' if self.lang == 'ar' else 'ar'
        # Save current page
        current = self.current_page if hasattr(self, 'current_page') else 'dashboard'
        # Rebuild UI
        for widget in self.root.winfo_children():
            widget.destroy()
        self.setup_ui()
        # Restore page
        if hasattr(self, 'switch_page'):
            self.switch_page(current)
    
    def open_features_dialog(self, display_label):
        """Open dialog to select multiple features"""
        dialog = tk.Toplevel(self.root)
        dialog.title(self.t('select_features'))
        dialog.geometry("450x550")
        dialog.configure(bg=COLORS['bg_card'])
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (450 // 2)
        y = (dialog.winfo_screenheight() // 2) - (550 // 2)
        dialog.geometry(f'450x550+{x}+{y}')
        
        # Header
        header = tk.Frame(dialog, bg=COLORS['accent_blue'], height=60)
        header.pack(fill='x')
        header.pack_propagate(False)
        
        tk.Label(
            header,
            text="🎯 " + self.t('select_features'),
            font=("Arial", 14, "bold"),
            bg=COLORS['accent_blue'],
            fg=COLORS['text_primary']
        ).pack(pady=18)
        
        # Content
        content = tk.Frame(dialog, bg=COLORS['bg_card'])
        content.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Scrollable frame
        canvas = tk.Canvas(content, bg=COLORS['bg_card'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(content, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=COLORS['bg_card'])
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Features list
        feature_options = [
            'night_vision', 'motion_detection', 'audio', 'ptz',
            'waterproof', 'vandal_proof', 'face_recognition',
            'license_plate', 'wide_angle', 'zoom', 'two_way_audio', 'smart_tracking'
        ]
        
        checkboxes = {}
        checkbox_vars = {}
        
        for feature in feature_options:
            var = tk.BooleanVar(value=feature in self.selected_features)
            checkbox_vars[feature] = var
            
            cb_frame = tk.Frame(scrollable_frame, bg=COLORS['bg_main'], 
                               highlightthickness=1, highlightbackground=COLORS['border'])
            cb_frame.pack(fill='x', pady=5, padx=5)
            
            cb = tk.Checkbutton(
                cb_frame,
                text=self.t(feature),
                variable=var,
                font=("Arial", 11),
                bg=COLORS['bg_main'],
                fg=COLORS['text_secondary'],
                selectcolor=COLORS['bg_input'],
                activebackground=COLORS['bg_main'],
                activeforeground=COLORS['text_primary'],
                cursor='hand2',
                padx=15,
                pady=10
            )
            cb.pack(fill='x')
            checkboxes[feature] = cb
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Buttons
        button_frame = tk.Frame(dialog, bg=COLORS['bg_card'])
        button_frame.pack(fill='x', padx=20, pady=(0, 20))
        
        def save_selection():
            self.selected_features = [f for f, v in checkbox_vars.items() if v.get()]
            
            # Update display
            if self.selected_features:
                display_text = ", ".join([self.t(f) for f in self.selected_features[:2]])
                if len(self.selected_features) > 2:
                    display_text += f" +{len(self.selected_features) - 2}"
                display_label.config(text=display_text, fg=COLORS['text_primary'])
            else:
                display_label.config(text=self.t('select_features'), fg=COLORS['text_dim'])
            
            dialog.destroy()
        
        def cancel():
            dialog.destroy()
        
        tk.Button(
            button_frame,
            text="✅ " + ("حفظ" if self.lang == 'ar' else "Save"),
            command=save_selection,
            bg=COLORS['accent_green'],
            fg=COLORS['text_primary'],
            font=("Arial", 12, "bold"),
            relief='flat',
            cursor='hand2',
            padx=20,
            pady=10,
            borderwidth=0
        ).pack(side='left', fill='x', expand=True, padx=(0, 5))
        
        tk.Button(
            button_frame,
            text="❌ " + ("إلغاء" if self.lang == 'ar' else "Cancel"),
            command=cancel,
            bg=COLORS['border'],
            fg=COLORS['text_primary'],
            font=("Arial", 12, "bold"),
            relief='flat',
            cursor='hand2',
            padx=20,
            pady=10,
            borderwidth=0
        ).pack(side='right', fill='x', expand=True, padx=(5, 0))
    
    def upload_image(self):
        """Upload image from device"""
        file_path = filedialog.askopenfilename(
            title="اختر صورة" if self.lang == 'ar' else "Choose Image",
            filetypes=[
                ("Image files", "*.png *.jpg *.jpeg *.gif *.bmp"),
                ("All files", "*.*")
            ]
        )
        
        if file_path:
            try:
                # Open and resize image
                image = Image.open(file_path)
                
                # Convert to RGB if necessary
                if image.mode in ('RGBA', 'LA', 'P'):
                    background = Image.new('RGB', image.size, (255, 255, 255))
                    if image.mode == 'P':
                        image = image.convert('RGBA')
                    background.paste(image, mask=image.split()[-1] if image.mode in ('RGBA', 'LA') else None)
                    image = background
                elif image.mode != 'RGB':
                    image = image.convert('RGB')
                
                # Create thumbnail for display
                display_img = image.copy()
                display_img.thumbnail((300, 200))
                
                # Save original to base64
                buffered = io.BytesIO()
                image.save(buffered, format="JPEG", quality=85)
                img_str = base64.b64encode(buffered.getvalue()).decode()
                self.current_screenshot = img_str
                
                # Display thumbnail
                photo = ImageTk.PhotoImage(display_img)
                self.screenshot_label.configure(image=photo, text="")
                self.screenshot_label.image = photo
                
                messagebox.showinfo(
                    "نجح" if self.lang == 'ar' else "Success",
                    "تم رفع الصورة بنجاح!" if self.lang == 'ar' else "Image uploaded successfully!"
                )
            except Exception as e:
                messagebox.showerror(
                    "خطأ" if self.lang == 'ar' else "Error",
                    f"فشل رفع الصورة: {str(e)}" if self.lang == 'ar' 
                    else f"Failed to upload image: {str(e)}"
                )
    
    def take_screenshot(self):
        """Legacy method - now redirects to upload_image"""
        self.upload_image()
    
    def add_camera(self):
        camera_data = {}
        for field, entry in self.entries.items():
            if field == 'features':
                # Save features as comma-separated string
                camera_data[field] = ','.join(self.selected_features) if self.selected_features else ''
            elif isinstance(entry, ttk.Combobox):
                value = entry.get()
                for key, trans in TRANSLATIONS[self.lang].items():
                    if trans == value:
                        camera_data[field] = key
                        break
                else:
                    camera_data[field] = value
            elif isinstance(entry, tk.Label):
                # Skip label entries (features display)
                camera_data[field] = ','.join(self.selected_features) if field == 'features' else ''
            else:
                camera_data[field] = entry.get()
        
        if not camera_data['camera_number'] or not camera_data['device_ip']:
            messagebox.showwarning(
                "تحذير" if self.lang == 'ar' else "Warning",
                "رقم الكاميرا و IP مطلوبان!" if self.lang == 'ar' 
                else "Camera number and IP are required!"
            )
            return
        
        camera_data['screenshot'] = self.current_screenshot
        camera_data['created_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        self.cameras_data.append(camera_data)
        self.save_data()
        self.clear_form()
        
        messagebox.showinfo(
            "نجح" if self.lang == 'ar' else "Success",
            "تمت إضافة الكاميرا بنجاح!" if self.lang == 'ar' 
            else "Camera added successfully!"
        )
    
    def update_camera(self):
        if not hasattr(self, 'tree') or not self.tree.selection():
            messagebox.showwarning(
                "تحذير" if self.lang == 'ar' else "Warning",
                "اختر كاميرا من قائمة الكاميرات أولاً!" if self.lang == 'ar' 
                else "Select a camera from the cameras list first!"
            )
            return
        
        index = self.tree.index(self.tree.selection()[0])
        camera_data = {}
        
        for field, entry in self.entries.items():
            if field == 'features':
                camera_data[field] = ','.join(self.selected_features) if self.selected_features else ''
            elif isinstance(entry, ttk.Combobox):
                value = entry.get()
                for key, trans in TRANSLATIONS[self.lang].items():
                    if trans == value:
                        camera_data[field] = key
                        break
                else:
                    camera_data[field] = value
            elif isinstance(entry, tk.Label):
                camera_data[field] = ','.join(self.selected_features) if field == 'features' else ''
            else:
                camera_data[field] = entry.get()
        
        camera_data['screenshot'] = self.current_screenshot or self.cameras_data[index].get('screenshot')
        camera_data['updated_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        camera_data['created_at'] = self.cameras_data[index].get('created_at', '')
        
        self.cameras_data[index] = camera_data
        self.save_data()
        self.clear_form()
        
        messagebox.showinfo(
            "نجح" if self.lang == 'ar' else "Success",
            "تم التحديث بنجاح!" if self.lang == 'ar' else "Updated successfully!"
        )
    
    def delete_camera(self):
        # Check if we have entries (we're on add_camera page)
        if hasattr(self, 'entries') and self.entries:
            # Try to find camera by camera_number from form
            camera_number = self.entries.get('camera_number')
            if camera_number and camera_number.get():
                cam_num = camera_number.get()
                for idx, camera in enumerate(self.cameras_data):
                    if camera.get('camera_number') == cam_num:
                        if messagebox.askyesno(
                            "تأكيد" if self.lang == 'ar' else "Confirm",
                            f"حذف الكاميرا {cam_num}؟" if self.lang == 'ar' else f"Delete camera {cam_num}?"
                        ):
                            del self.cameras_data[idx]
                            self.save_data()
                            self.clear_form()
                            messagebox.showinfo(
                                "نجح" if self.lang == 'ar' else "Success",
                                "تم الحذف!" if self.lang == 'ar' else "Deleted!"
                            )
                        return
        
        # Otherwise check if we have tree selection (we're on cameras page)
        if hasattr(self, 'tree') and self.tree.selection():
            if messagebox.askyesno(
                "تأكيد" if self.lang == 'ar' else "Confirm",
                "حذف هذه الكاميرا؟" if self.lang == 'ar' else "Delete this camera?"
            ):
                index = self.tree.index(self.tree.selection()[0])
                del self.cameras_data[index]
                self.save_data()
                self.refresh_table()
                messagebox.showinfo(
                    "نجح" if self.lang == 'ar' else "Success",
                    "تم الحذف!" if self.lang == 'ar' else "Deleted!"
                )
            return
        
        # No selection anywhere
        messagebox.showwarning(
            "تحذير" if self.lang == 'ar' else "Warning",
            "اختر كاميرا للحذف أو أدخل رقم الكاميرا!" if self.lang == 'ar' 
            else "Select a camera to delete or enter camera number!"
        )
    
    def clear_form(self):
        for field, entry in self.entries.items():
            if field == 'features':
                self.selected_features = []
                if isinstance(entry, tk.Label):
                    entry.config(text=self.t('select_features'), fg=COLORS['text_dim'])
            elif isinstance(entry, ttk.Combobox):
                entry.set('')
            else:
                if hasattr(entry, 'delete'):
                    entry.delete(0, tk.END)
        
        self.current_screenshot = None
        if hasattr(self, 'screenshot_label'):
            self.screenshot_label.configure(
                image='',
                text="🖼️\n" + self.t('upload_image')
            )
    
    def on_row_double_click(self, event):
        selected = self.tree.selection()
        if selected:
            index = self.tree.index(selected[0])
            camera = self.cameras_data[index]
            
            self.switch_page('add_camera')
            self.root.after(100, lambda: self.fill_form(camera))
    
    def fill_form(self, camera):
        for field, entry in self.entries.items():
            value = camera.get(field, '')
            
            if field == 'features':
                # Load features from comma-separated string
                if value:
                    self.selected_features = value.split(',')
                    display_text = ", ".join([self.t(f) for f in self.selected_features[:2]])
                    if len(self.selected_features) > 2:
                        display_text += f" +{len(self.selected_features) - 2}"
                    if isinstance(entry, tk.Label):
                        entry.config(text=display_text, fg=COLORS['text_primary'])
                else:
                    self.selected_features = []
                    if isinstance(entry, tk.Label):
                        entry.config(text=self.t('select_features'), fg=COLORS['text_dim'])
            elif isinstance(entry, ttk.Combobox):
                translated = self.t(value) if value else ''
                entry.set(translated)
            else:
                if hasattr(entry, 'delete') and hasattr(entry, 'insert'):
                    entry.delete(0, tk.END)
                    entry.insert(0, value)
        
        if camera.get('screenshot'):
            try:
                img_data = base64.b64decode(camera['screenshot'])
                image = Image.open(io.BytesIO(img_data))
                image.thumbnail((300, 200))
                photo = ImageTk.PhotoImage(image)
                self.screenshot_label.configure(image=photo, text="")
                self.screenshot_label.image = photo
                self.current_screenshot = camera['screenshot']
            except:
                pass
    
    def refresh_table(self):
        if not hasattr(self, 'tree'):
            return
        
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        for camera in self.cameras_data:
            values = (
                camera.get('camera_number', ''),
                camera.get('branch_name', ''),
                camera.get('branch_code', ''),
                camera.get('device_ip', ''),
                camera.get('location', ''),
                self.t(camera.get('location_type', '')),
                self.t(camera.get('direction', '')),
                self.t(camera.get('status', ''))
            )
            self.tree.insert('', 'end', values=values)
    
    def search_cameras(self):
        if not hasattr(self, 'tree'):
            return
        
        search_term = self.search_var.get().lower()
        
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        for camera in self.cameras_data:
            if any(search_term in str(value).lower() for value in camera.values()):
                values = (
                    camera.get('camera_number', ''),
                    camera.get('branch_name', ''),
                    camera.get('branch_code', ''),
                    camera.get('device_ip', ''),
                    camera.get('location', ''),
                    self.t(camera.get('location_type', '')),
                    self.t(camera.get('direction', '')),
                    self.t(camera.get('status', ''))
                )
                self.tree.insert('', 'end', values=values)
    
    def export_to_excel(self):
        if not self.cameras_data:
            messagebox.showwarning(
                "تحذير" if self.lang == 'ar' else "Warning",
                "لا توجد بيانات للتصدير!" if self.lang == 'ar' 
                else "No data to export!"
            )
            return
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"cameras_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if file_path:
            try:
                export_data = []
                for camera in self.cameras_data:
                    row = {}
                    for field in ['camera_number', 'branch_name', 'branch_code', 'device_ip',
                                 'dvr_name', 'username_field', 'password_field', 'location',
                                 'location_type', 'direction', 'working_hours', 'model',
                                 'serial_number', 'features', 'lens_layout',
                                 'rtsp_url', 'status', 'notes']:
                        value = camera.get(field, '')
                        if field in ['location_type', 'direction', 'working_hours', 'status']:
                            value = self.t(value) if value else ''
                        row[self.t(field)] = value
                    export_data.append(row)
                
                df = pd.DataFrame(export_data)
                
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Cameras', index=False)
                    
                    workbook = writer.book
                    worksheet = writer.sheets['Cameras']
                    
                    # Styling
                    header_fill = PatternFill(start_color="2E69FF", end_color="2E69FF", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF", size=11)
                    cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    thin_border = Border(
                        left=Side(style='thin', color='636466'),
                        right=Side(style='thin', color='636466'),
                        top=Side(style='thin', color='636466'),
                        bottom=Side(style='thin', color='636466')
                    )
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = cell_alignment
                        cell.border = thin_border
                    
                    for col_idx, column in enumerate(df.columns, 1):
                        max_length = len(str(column))
                        for row_idx, value in enumerate(df[column], 2):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.alignment = cell_alignment
                            cell.border = thin_border
                            cell.font = Font(size=10, color="282829")
                            max_length = max(max_length, len(str(value)))
                        
                        worksheet.column_dimensions[chr(64 + col_idx)].width = min(max_length + 3, 50)
                    
                    # Screenshots
                    screenshot_col = len(df.columns) + 1
                    worksheet.cell(row=1, column=screenshot_col).value = self.t('screenshot')
                    worksheet.cell(row=1, column=screenshot_col).fill = header_fill
                    worksheet.cell(row=1, column=screenshot_col).font = header_font
                    worksheet.cell(row=1, column=screenshot_col).alignment = cell_alignment
                    worksheet.cell(row=1, column=screenshot_col).border = thin_border
                    worksheet.column_dimensions[chr(64 + screenshot_col)].width = 28
                    
                    for idx, camera in enumerate(self.cameras_data, 2):
                        if camera.get('screenshot'):
                            try:
                                img_data = base64.b64decode(camera['screenshot'])
                                img = Image.open(io.BytesIO(img_data))
                                img.thumbnail((160, 110))
                                
                                img_byte_arr = io.BytesIO()
                                img.save(img_byte_arr, format='PNG')
                                img_byte_arr.seek(0)
                                
                                xl_img = XLImage(img_byte_arr)
                                xl_img.width = 160
                                xl_img.height = 110
                                
                                cell_ref = f"{chr(64 + screenshot_col)}{idx}"
                                worksheet.add_image(xl_img, cell_ref)
                                worksheet.row_dimensions[idx].height = 85
                            except Exception as e:
                                print(f"Error adding image: {e}")
                
                messagebox.showinfo(
                    "نجح" if self.lang == 'ar' else "Success",
                    f"تم التصدير بنجاح!\n{file_path}" if self.lang == 'ar' 
                    else f"Exported successfully!\n{file_path}"
                )
            
            except Exception as e:
                messagebox.showerror(
                    "خطأ" if self.lang == 'ar' else "Error",
                    f"فشل التصدير: {str(e)}" if self.lang == 'ar' 
                    else f"Export failed: {str(e)}"
                )
    
    def import_from_excel(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")],
            title="اختر ملف Excel" if self.lang == 'ar' else "Select Excel File"
        )
        
        if file_path:
            try:
                df = pd.read_excel(file_path)
                
                reverse_map = {}
                for key in ['camera_number', 'branch_name', 'branch_code', 'device_ip',
                           'dvr_name', 'username_field', 'password_field', 'location',
                           'location_type', 'direction', 'working_hours', 'model',
                           'serial_number', 'features', 'lens_layout',
                           'rtsp_url', 'status', 'notes']:
                    reverse_map[self.t(key)] = key
                
                imported_count = 0
                for _, row in df.iterrows():
                    camera_data = {}
                    for col_name, value in row.items():
                        if col_name in reverse_map:
                            field = reverse_map[col_name]
                            
                            if field in ['location_type', 'direction', 'working_hours', 'status']:
                                for key, trans in TRANSLATIONS[self.lang].items():
                                    if trans == str(value):
                                        camera_data[field] = key
                                        break
                                else:
                                    camera_data[field] = str(value) if pd.notna(value) else ''
                            else:
                                camera_data[field] = str(value) if pd.notna(value) else ''
                    
                    if camera_data.get('camera_number'):
                        camera_data['created_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        self.cameras_data.append(camera_data)
                        imported_count += 1
                
                self.save_data()
                
                messagebox.showinfo(
                    "نجح" if self.lang == 'ar' else "Success",
                    f"تم استيراد {imported_count} كاميرا!" if self.lang == 'ar' 
                    else f"Imported {imported_count} cameras!"
                )
            
            except Exception as e:
                messagebox.showerror(
                    "خطأ" if self.lang == 'ar' else "Error",
                    f"فشل الاستيراد: {str(e)}" if self.lang == 'ar' 
                    else f"Import failed: {str(e)}"
                )
    
    def save_data(self):
        try:
            with open(self.data_file, 'w', encoding='utf-8') as f:
                json.dump(self.cameras_data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror(
                "خطأ" if self.lang == 'ar' else "Error",
                f"فشل الحفظ: {str(e)}" if self.lang == 'ar' else f"Save failed: {str(e)}"
            )
    
    def load_data(self):
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    self.cameras_data = json.load(f)
            except Exception as e:
                messagebox.showerror(
                    "خطأ" if self.lang == 'ar' else "Error",
                    f"فشل التحميل: {str(e)}" if self.lang == 'ar' else f"Load failed: {str(e)}"
                )
                self.cameras_data = []
    
    def logout(self):
        if messagebox.askyesno(
            "تأكيد" if self.lang == 'ar' else "Confirm",
            "تسجيل الخروج؟" if self.lang == 'ar' else "Logout?"
        ):
            self.root.destroy()
            main()

# ==================== MAIN ====================
def main():
    root = tk.Tk()
    root.withdraw()
    
    def on_login_success(username, lang):
        root.destroy()
        app_root = tk.Tk()
        app = DashboardApp(app_root, username, lang)
        app_root.mainloop()
    
    login_root = tk.Tk()
    LoginWindow(login_root, on_login_success)
    login_root.mainloop()

if __name__ == "__main__":
    main()